#!/usr/bin/env python3
"""
import-zamer.py

Export investorskych podkladu z klientskeho Excelu (Zamer VPD1) do JSON
pro pod-web /investori/. Zdroj pravdy = xlsx; tento skript se spousti
znovu pri kazde aktualizaci tabulky od klienta:

    uv run --with openpyxl scripts/import-zamer.py /cesta/k/zamer-vpd1.xlsx

Vystup: src/data/investori/*.json

Principy:
  - Slovni formulace, ciselne hodnoty i vzorce (propojeni bunek) se prenasi
    1:1, nikdy se neprepisuji. Vzorce zustavaji v JSON pro budouci
    interaktivni kalkulace.
  - Kazda hodnota nese: d = display string (presne dle Excel number formatu,
    cesky zapis: mezera tisice, carka desetiny), v = raw hodnota,
    f = vzorec (pokud existuje), h = hyperlink (pokud existuje).
  - Bunky s in-cell obrazky (rich data #VALUE!) se mapuji na soubory
    v public/images/investori/ podle IMG_MAP nize.
"""
import json
import re
import sys
import datetime
import zipfile
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / 'src' / 'data' / 'investori'

NBSP = ' '


# ---------------------------------------------------------------------------
# Format engine: Excel number_format -> cesky display string
# ---------------------------------------------------------------------------

def _czech_number(value, decimals, thousands=True):
    q = Decimal(str(value)).quantize(
        Decimal(1).scaleb(-decimals) if decimals else Decimal(1),
        rounding=ROUND_HALF_UP,
    )
    sign = '-' if q < 0 else ''
    q = abs(q)
    intpart, _, frac = str(q).partition('.')
    if thousands:
        groups = []
        while len(intpart) > 3:
            groups.insert(0, intpart[-3:])
            intpart = intpart[:-3]
        groups.insert(0, intpart)
        intpart = NBSP.join(groups)
    out = sign + intpart
    if decimals:
        out += ',' + (frac or '').ljust(decimals, '0')
    return out


def _parse_literals(nf):
    """Extract prefix/suffix literals around the numeric token."""
    # Strip color/conditional sections; keep first section only.
    section = nf.split(';')[0]
    # Tokenize: quoted literals, escaped chars, numeric core, percent.
    prefix, suffix = '', ''
    core_seen = False
    i = 0
    while i < len(section):
        ch = section[i]
        if ch == '"':
            j = section.index('"', i + 1)
            lit = section[i + 1:j]
            if core_seen:
                suffix += lit
            else:
                prefix += lit
            i = j + 1
        elif ch == '\\':
            lit = section[i + 1] if i + 1 < len(section) else ''
            if core_seen:
                suffix += lit
            else:
                prefix += lit
            i += 2
        elif ch in '#0.,?':
            core_seen = True
            i += 1
        elif ch == '%':
            suffix += '%'
            core_seen = True
            i += 1
        elif ch == '@':
            core_seen = True
            i += 1
        else:
            # raw literal char (space etc.)
            if core_seen:
                suffix += ch
            else:
                prefix += ch
            i += 1
    return prefix, suffix


def _numeric_core(nf):
    section = nf.split(';')[0]
    # remove quoted literals and escapes
    section = re.sub(r'"[^"]*"', '', section)
    section = re.sub(r'\\.', '', section)
    m = re.search(r'[#0][#0,]*(?:\.(0+))?', section)
    decimals = len(m.group(1)) if (m and m.group(1)) else 0
    thousands = bool(m and ',' in m.group(0))
    return decimals, thousands


def fmt(value, nf):
    """Display string presne dle Excel number formatu, cesky."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime.datetime, datetime.date)):
        return f'{value.day}. {value.month}. {value.year}'
    if isinstance(value, bool):
        return 'ano' if value else 'ne'
    nf = nf or 'General'
    if nf in ('General', '@'):
        # Excel General nikdy neseskupuje tisice (roky, parcelni cisla)
        # a nezaokrouhluje; drzime plnou presnost hodnoty.
        if isinstance(value, float) and not value.is_integer():
            return f'{value:.10g}'.replace('.', ',')
        return _czech_number(value, 0, thousands=False)
    is_percent = '%' in re.sub(r'"[^"]*"', '', nf)
    decimals, thousands = _numeric_core(nf)
    prefix, suffix = _parse_literals(nf)
    v = value * 100 if is_percent else value
    return prefix + _czech_number(v, decimals, thousands) + suffix


# ---------------------------------------------------------------------------
# In-cell obrazky (rich data): bunky se v openpyxl ukazou jako '#VALUE!'.
# Nektera z nich nesou i text (napr. badge "terasa"/"balkon" u Klecanske
# aleje); mapujeme cell -> text, ostatni image bunky renderujeme prazdne.
# ---------------------------------------------------------------------------

def build_rich_text_map(xlsx_path):
    z = zipfile.ZipFile(xlsx_path)
    names = set(z.namelist())
    if 'xl/richData/rdrichvalue.xml' not in names:
        return {}
    rdrv = z.read('xl/richData/rdrichvalue.xml').decode()
    rvs = re.findall(r'<rv s="(\d+)">(.*?)</rv>', rdrv, re.S)
    structures = re.findall(r'<s t="[^"]*">(.*?)</s>',
                            z.read('xl/richData/rdrichvaluestructure.xml').decode(), re.S)
    struct_keys = [re.findall(r'<k n="([^"]+)"', s) for s in structures]
    rv_text = []
    for s_idx, body in rvs:
        keys = struct_keys[int(s_idx)]
        vals = re.findall(r'<v>([^<]*)</v>', body)
        text = None
        for k, v in zip(keys, vals):
            if k == 'Text':
                text = v
        rv_text.append(text)
    wb_xml = z.read('xl/workbook.xml').decode()
    wb_rels = z.read('xl/_rels/workbook.xml.rels').decode()
    relmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', wb_rels))
    out = {}
    for sname, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml):
        sname = sname.replace('&amp;', '&')
        target = relmap.get(rid, '')
        path = f"xl/{target.lstrip('/')}" if not target.startswith('xl/') else target
        if path not in names:
            continue
        sx = z.read(path).decode()
        for ref, vm in re.findall(r'<c r="([A-Z]+\d+)"[^>]*vm="(\d+)"', sx):
            text = rv_text[int(vm) - 1] if 0 < int(vm) <= len(rv_text) else None
            out[(sname, ref)] = text or ''
    return out


# ---------------------------------------------------------------------------
# Cell accessor
# ---------------------------------------------------------------------------

class Sheet:
    def __init__(self, wsf, wsv, richmap=None):
        self.wsf = wsf  # formulas
        self.wsv = wsv  # cached values
        self.richmap = richmap or {}

    def cell(self, ref):
        cf = self.wsf[ref]
        cv = self.wsv[ref]
        rec = {}
        val = cv.value
        if isinstance(val, str):
            val = val.replace('\xa0', ' ')
        # In-cell obrazek: '#VALUE!' nahradime jeho textem (je-li), jinak ''.
        if val == '#VALUE!':
            val = self.richmap.get((self.wsf.title, ref), '')
        rec['d'] = fmt(val, cf.number_format)
        if isinstance(val, (datetime.datetime, datetime.date)):
            rec['v'] = val.isoformat()[:10]
        else:
            rec['v'] = val
        if isinstance(cf.value, str) and cf.value.startswith('='):
            rec['f'] = cf.value
        if cf.hyperlink and cf.hyperlink.target and str(cf.hyperlink.target).startswith('http'):
            rec['h'] = cf.hyperlink.target
        return rec

    def d(self, ref):
        return self.cell(ref)['d']

    def row(self, refs):
        return [self.cell(r) for r in refs]


def srow(sh, cols, n):
    return sh.row([f'{c}{n}' for c in cols])


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------

def main(xlsx_path):
    wbf = openpyxl.load_workbook(xlsx_path, data_only=False)
    wbv = openpyxl.load_workbook(xlsx_path, data_only=True)
    richmap = build_rich_text_map(xlsx_path)

    iz = Sheet(wbf['Investiční záměr'], wbv['Investiční záměr'], richmap)
    vy = Sheet(wbf['Základní údaje & výpočty'], wbv['Základní údaje & výpočty'], richmap)
    sc = Sheet(wbf['Základní scénáře'], wbv['Základní scénáře'], richmap)
    s1 = Sheet(wbf['S1 (base) '], wbv['S1 (base) '], richmap)
    ma = Sheet(wbf['Mapa Areálu'], wbv['Mapa Areálu'], richmap)
    po = Sheet(wbf['Poznámky'], wbv['Poznámky'], richmap)

    OUT.mkdir(parents=True, exist_ok=True)

    def write(name, data):
        path = OUT / name
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=1) + '\n',
            encoding='utf-8',
        )
        print(f'  {path.relative_to(REPO)}')

    # --- meta ---------------------------------------------------------------
    meta = {
        'verze': iz.d('D1'),
        'datum': iz.cell('D2')['v'],
        'datumDisplay': iz.d('D2'),
        'aktualizace': iz.d('B7'),
        'badge': iz.d('D9'),
        'zdroj': Path(xlsx_path).name,
    }
    write('meta.json', meta)

    # --- zamer.json (Investicni zamer) --------------------------------------
    def fact_rows(rows):
        out = []
        for n in rows:
            label = iz.d(f'B{n}')
            if not label:
                continue
            out.append({'label': label, **{k: v for k, v in iz.cell(f'D{n}').items()}})
        return out

    zamer = {
        'titul': iz.d('B5'),
        'podtitul': iz.d('B6'),
        'sekce': [
            {
                'id': 'aktivum',
                'titul': iz.d('B10'),
                'radky': fact_rows(range(11, 23)),
                'footnote': iz.d('B23'),
            },
            {
                'id': 'zakladni',
                'titul': iz.d('B35'),
                'radky': fact_rows(range(36, 58)),
                'footnote': iz.d('B58'),
            },
            {
                'id': 'ostatni',
                'titul': iz.d('B61'),
                'radky': fact_rows(range(62, 81)),
                'footnote': iz.d('B81'),
            },
        ],
    }
    write('zamer.json', zamer)

    # --- mapa.json (ortofotomapa legenda; stejna data B26:D29 i Mapa Arealu) -
    mapa = {
        'titul': iz.d('B26'),
        'zony': [
            {
                'id': 'areal',
                'vymera': iz.cell('B27'),
                'text': iz.d('C27'),
            },
            {
                'id': 'jadro',
                'vymera': iz.cell('B28'),
                'text': iz.d('C28'),
            },
            {
                'id': 'zazemi',
                'vymera': iz.cell('B29'),
                'text': iz.d('C29'),
            },
        ],
        'obrazek': '/images/investori/ortofotomapa.jpg',
    }
    write('mapa.json', mapa)

    # --- scenare.json --------------------------------------------------------
    headers = [sc.d(f'{c}5') for c in 'DEFGHIJK']
    scen_rows = []
    for n in range(6, 11):
        cells = {c: sc.cell(f'{c}{n}') for c in 'DEFGHIJK'}
        scen_rows.append({
            'nazev': cells['D']['d'],
            'krok': cells['E']['d'],
            'kod': cells['F']['d'],
            'detail': cells['G']['d'],
            'dejstvi': cells['H']['d'],
            'odkup': cells['I']['d'],
            'kapital': cells['J']['d'],
            'alternativa': cells['K']['d'],
        })
    scenare = {
        'titul': sc.d('D2'),
        'headers': headers,
        'scenare': scen_rows,
        'footnote': sc.d('D11'),
    }
    write('scenare.json', scenare)

    # --- s1.json (header investicne-realizacniho scenare S1 base) ------------
    s1_info = []
    for n in range(10, 30):
        label = s1.d(f'C{n}')
        if not label:
            continue
        s1_info.append({'label': label, **s1.cell(f'D{n}')})
    s1_struktura = []
    for n in range(67, 72):
        s1_struktura.append({
            'faze': s1.d(f'C{n}'),
            'obsah': s1.cell(f'D{n}'),
            'delka': s1.cell(f'E{n}'),
        })
    s1_scenar = []
    for n in range(40, 45):
        s1_scenar.append({
            'dejstvi': s1.d(f'C{n}'),
            'marketing': s1.d(f'E{n}'),
        })
    s1data = {
        'titul': s1.d('C3'),
        'podtitul': s1.d('C4'),
        'badge': s1.d('C5'),
        'zakladniInformace': {'titul': s1.d('C9'), 'radky': s1_info},
        'projekt': {
            'titul': s1.d('C34'),
            'popis': s1.d('C36'),
            'proUbytovane': s1.d('D36'),
            'proDevelopera': s1.d('G36'),
        },
        'scenar': {
            'titul': s1.d('C39'),
            'marketingHeader': s1.d('E39'),
            'dejstvi': s1_scenar,
            'footnote': s1.d('C45'),
        },
        'model': {
            'titul': s1.d('C49'),
            'infoTitul': s1.d('C54'),
            'popis': s1.d('C56'),
            'strukturaTitul': s1.d('C66'),
            'strukturaHeaders': [s1.d('D66'), s1.d('E66')],
            'struktura': s1_struktura,
        },
    }
    write('s1.json', s1data)

    # --- poznamky.json --------------------------------------------------------
    polozky = []
    for n in range(8, 107):
        text = po.d(f'C{n}')
        if not text:
            continue
        polozky.append({'cislo': po.d(f'B{n}'), 'text': text})
    write('poznamky.json', {'titul': po.d('B3'), 'polozky': polozky})

    # --- vypocty.json ----------------------------------------------------------
    # Radky parcel: 8..36 (tabulky 1 a 2 sdili radky pres cislo parcely).
    PARCELY_COLS = list('CDEFGH') + list('LMNOPQRSTUVWX')
    parcely_rows = []
    for n in range(8, 37):
        if vy.d(f'C{n}') == '':
            continue
        parcely_rows.append(srow(vy, PARCELY_COLS, n))
    parcely = {
        'titul1': vy.d('B4'),
        'titul2': vy.d('K4'),
        'headers': [vy.d(f'{c}7') for c in PARCELY_COLS],
        'rows': parcely_rows,
        'souctyLabels': {c: vy.d(f'{c}38') for c in 'EHMNOSTUVWX' if vy.d(f'{c}38')},
        'soucty': {c: vy.cell(f'{c}39') for c in 'EHMNOSTUVWX'},
        'poznamka': vy.d('N40'),
    }

    zony = {
        'jadro': {
            'titul': vy.d('B42'),
            'zatizeneLabel': vy.d('C42'),
            'vymera': vy.cell('C43'),
            'podilArealLabel': vy.d('E42'),
            'podilAreal': vy.cell('E43'),
            'podilSMLabel': vy.d('E44'),
            'podilSM': vy.cell('E45'),
        },
        'celek': {
            'titul': vy.d('B69'),
            'zatizeneLabel': vy.d('C69'),
            'vymera': vy.cell('C70'),
            'podilArealLabel': vy.d('E69'),
            'podilAreal': vy.cell('E70'),
            'podilSMLabel': vy.d('E71'),
            'podilSM': vy.cell('E72'),
        },
        'zazemi': {
            'titul': vy.d('B118'),
            'zatizeneLabel': vy.d('C118'),
            'vymera': vy.cell('C119'),
            'podilArealLabel': vy.d('E118'),
            'podilAreal': vy.cell('E119'),
            'podilSMLabel': vy.d('E120'),
            'podilSM': vy.cell('E121'),
        },
    }

    zalohy = {
        'titul': vy.d('AC4'),
        'headers': [vy.d(f'{c}7') for c in ['AC', 'AD', 'AE', 'AF', 'AG']],
        'rows': [srow(vy, ['AC', 'AD', 'AE', 'AF', 'AG'], n) for n in range(8, 24)],
        'mimoradna': {
            'titul': vy.d('AC26'),
            'popis': vy.d('AC27'),
            'vyse': vy.cell('AF27'),
            'celkem': vy.cell('AG27'),
        },
    }

    KC_COLS = ['AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX']
    kupni_cena = {
        'titul': vy.d('AJ4'),
        'params': [
            {'label': vy.d('AJ5'), **vy.cell('AJ6')},
            {'label': vy.d('AN5'), **vy.cell('AN6')},
            {'label': vy.d('AS5'), **vy.cell('AS6')},
            {'label': vy.d('AU5'), **vy.cell('AU6')},
        ],
        'headers': [vy.d(f'{c}7') for c in KC_COLS[:13]] + [vy.d('AW5'), vy.d('AX5')],
        'rows': [srow(vy, KC_COLS, n) for n in range(8, 24)],
        'footnote': vy.d('AJ24'),
        'note': vy.d('AJ28'),
    }

    NAJEM_COLS = ['BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ']
    najem_rows = []
    for n in range(8, 37):
        if vy.d(f'BI{n}') == '':
            continue
        najem_rows.append(srow(vy, NAJEM_COLS, n))
    najem = {
        'titul': vy.d('BI4'),
        'params': [
            {'label': vy.d('BO5'), **vy.cell('BO6')},
            {'label': vy.d('BQ5'), **vy.cell('BQ6')},
        ],
        'headers': [vy.d('BI5'), vy.d('BJ5'), vy.d('BK5'),
                    vy.d('BL7'), vy.d('BM7'), vy.d('BN7'),
                    vy.d('BO7'), vy.d('BP7'), vy.d('BQ7')],
        'podlaziHeader': vy.d('BL5'),
        'potencialLabel': vy.d('BL38'),
        'potencial': vy.cell('BL39'),
        'rows': najem_rows,
        'souctyLabels': {c: vy.d(f'{c}38') for c in ['BO', 'BP', 'BQ']},
        'soucty': {c: vy.cell(f'{c}39') for c in ['BO', 'BP', 'BQ']},
        'refCaption': vy.d('BI43'),
        'refRows': [
            [vy.cell('BO41'), vy.cell('BP41'), vy.cell('BQ41')],
            [vy.cell('BO42'), vy.cell('BP42'), vy.cell('BQ42')],
        ],
    }

    rust_prijmu = {
        'titul': vy.d('BS4'),
        'params': [
            {'label': vy.d('BS5'), **vy.cell('BS6')},
            {'label': vy.d('BU5'), **vy.cell('BU6')},
        ],
        'headers': [vy.d('BS7'), vy.d('BT7'), vy.d('BU7'), vy.d('BV5')],
        'rows': [srow(vy, ['BS', 'BT', 'BU', 'BV'], n) for n in range(8, 24)],
        'totalLabel': vy.d('BV36'),
        'total': vy.cell('BV39'),
    }

    porovnani = {
        'titul': vy.d('BY4'),
        'headers': [vy.d(f'{c}5') for c in ['BY', 'BZ', 'CA', 'CB', 'CC']],
        'rows': [srow(vy, ['BY', 'BZ', 'CA', 'CB', 'CC'], n) for n in range(8, 24)],
        'totals': [
            {'label': vy.d('BZ36'), **vy.cell('BZ39')},
            {'label': vy.d('CB36'), **vy.cell('CB39')},
        ],
    }

    TC_COLS = ['CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM']
    tc_rows = []
    for n in range(8, 37):
        if vy.d(f'CG{n}') == '':
            continue
        tc_rows.append(srow(vy, TC_COLS, n))
    trzni_cena = {
        'titul': vy.d('CG4'),
        'param': {'label': vy.d('CL5'), **vy.cell('CL6')},
        'headers': [vy.d('CG5'), vy.d('CH5'), vy.d('CI5'), vy.d('CJ5'), vy.d('CK5'),
                    vy.d('CL7'), vy.d('CM7')],
        'rows': tc_rows,
        'totals': [
            {'label': vy.d('CL38'), **vy.cell('CL39')},
            {'label': vy.d('CM38'), **vy.cell('CM39')},
        ],
        'refCaption': vy.d('CG43'),
        'refRows': [[vy.cell('CK42'), vy.cell('CL42'), vy.cell('CM42')]],
    }

    trzni_rust = {
        'titul': vy.d('CP4'),
        'params': [
            {'label': vy.d('CP5'), **vy.cell('CP6')},
            {'label': vy.d('CR5'), **vy.cell('CR6')},
        ],
        'headers': [vy.d('CP7'), vy.d('CQ7'), vy.d('CR7')],
        'rows': [srow(vy, ['CP', 'CQ', 'CR'], n) for n in range(8, 24)],
        'totals': [
            {'label': vy.d('CQ38'), **vy.cell('CQ39')},
            {'label': vy.d('CR38'), **vy.cell('CR39')},
        ],
    }

    rozdil = {
        'titul': vy.d('CU4'),
        'param': {'label': vy.d('CU5'), **vy.cell('CU6')},
        'param2': {'label': vy.d('CX5'), **vy.cell('CX6')},
        'headers': [vy.d(f'{c}7') for c in ['CU', 'CV', 'CW', 'CX', 'CY', 'CZ']],
        'rows': [srow(vy, ['CU', 'CV', 'CW', 'CX', 'CY', 'CZ'], n) for n in range(8, 24)],
        'totalLabel': vy.d('CX36'),
        'total': vy.cell('CX39'),
    }

    UH_COLS = ['DM', 'DN', 'DO', 'DP', 'DQ', 'DR', 'DT', 'DU', 'DV']
    uh_rows = []
    for n in range(8, 37):
        if vy.d(f'DM{n}') == '':
            continue
        uh_rows.append(srow(vy, UH_COLS, n))
    ubytovaci_hub = {
        'titul': vy.d('DM4'),
        'params': [
            {'label': vy.d('DP5'), **vy.cell('DP6')},
            {'label': vy.d('DR5'), **vy.cell('DR6')},
            {'label': vy.d('DT5'), **vy.cell('DT6')},
            {'label': vy.d('DU5'), **vy.cell('DU6')},
            {'label': vy.d('DV5'), **vy.cell('DV6')},
        ],
        'headers': [vy.d('DM5'), vy.d('DN5'), vy.d('DO5'),
                    vy.d('DP7'), vy.d('DQ7'), vy.d('DR7'),
                    vy.d('DT7'), vy.d('DU7'), vy.d('DV7')],
        'rows': uh_rows,
        'souctyLabels': {c: vy.d(f'{c}38') for c in ['DR', 'DT', 'DU', 'DV']},
        'soucty': {c: vy.cell(f'{c}39') for c in ['DR', 'DT', 'DU', 'DV']},
        'refCaption': vy.d('DM43'),
    }

    NC_COLS = ['EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ', 'ER', 'ES']
    nova_ctvrt = {
        'titul': vy.d('EH4'),
        'params': [
            {'label': vy.d('EH5'), **vy.cell('EH6')},
            {'label': vy.d('EJ5'), **vy.cell('EJ6')},
            {'label': vy.d('EK5'), **vy.cell('EK6')},
            {'label': vy.d('EL5'), **vy.cell('EL6')},
            {'label': vy.d('EM5'), **vy.cell('EM6')},
            {'label': vy.d('EN5'), **vy.cell('EN6')},
            {'label': vy.d('EO5'), **vy.cell('EO6')},
            {'label': vy.d('EP5'), **vy.cell('EP6')},
            {'label': vy.d('EQ5'), **vy.cell('EQ6')},
            {'label': vy.d('ER5'), **vy.cell('ER6')},
        ],
        'headers': [vy.d(f'{c}7') for c in NC_COLS],
        'rows': [srow(vy, NC_COLS, n) for n in range(8, 24)],
        'vysledekLabel': vy.d('EQ33'),
        'vysledek': vy.cell('EQ39'),
        'prumCenaLabel': vy.d('EM41'),
        'prumCena': vy.cell('EN41'),
        'caption': vy.d('EH42'),
    }

    CRE_COLS = ['EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN']
    cre_rows = [srow(vy, CRE_COLS, n) for n in range(188, 215)]
    creditas = {
        'titul': vy.d('EH185'),
        'prumLabel': vy.d('EN185'),
        'prum': vy.cell('EN186'),
        'headers': [vy.cell(f'{c}187') for c in CRE_COLS],
        'rows': cre_rows,
        'disclaimer': vy.d('EH215'),
    }

    vizualizace = [
        {'id': 'interier-1kk', 'titul': vy.d('BB4'), 'img': '/images/investori/vizualizace-interier-1kk.jpg'},
        {'id': 'exterier-jadro', 'titul': vy.d('BB41'), 'img': '/images/investori/vizualizace-exterier-jadro.jpg'},
        {'id': 'budova-d', 'titul': vy.d('BB109'), 'img': '/images/investori/vizualizace-budova-d.jpg'},
        {'id': 'interier-sdileny', 'titul': vy.d('DF4'), 'img': '/images/investori/vizualizace-interier-sdileny-pokoj.jpg'},
        {'id': 'nova-ctvrt', 'titul': vy.d('EA4'), 'img': '/images/investori/vizualizace-nova-ctvrt-1.jpg'},
        {'id': 'nova-ctvrt-2', 'titul': vy.d('EA4'), 'img': '/images/investori/vizualizace-nova-ctvrt-2.jpg'},
        {'id': 'nova-ctvrt-3', 'titul': vy.d('EA4'), 'img': '/images/investori/vizualizace-nova-ctvrt-3.jpg'},
    ]

    vypocty = {
        'titul': vy.d('B2'),
        'skupiny': {
            'areal': vy.d('B4') and 'Areál',
            'smlouvy': vy.d('AC2'),
            'startovaciHub': vy.d('BB2'),
            'ubytovaciHub': vy.d('DF2'),
            'novaCtvrt': vy.d('EA2'),
        },
        'mapaTitul': vy.d('B7'),
        'parcely': parcely,
        'zony': zony,
        'zalohy': zalohy,
        'kupniCena': kupni_cena,
        'najem': najem,
        'rustPrijmu': rust_prijmu,
        'porovnani': porovnani,
        'trzniCena': trzni_cena,
        'trzniRust': trzni_rust,
        'rozdil': rozdil,
        'ubytovaciHub': ubytovaci_hub,
        'novaCtvrt': nova_ctvrt,
        'creditas': creditas,
        'vizualizace': vizualizace,
    }
    write('vypocty.json', vypocty)

    print('OK')


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Pouziti: uv run --with openpyxl scripts/import-zamer.py <zamer-vpd1.xlsx>')
        sys.exit(1)
    main(sys.argv[1])
